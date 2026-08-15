from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMNS = [
    ("Адрес", "address", 40),
    ("Кол-во бутылей", "quantity", 16),
    ("Сумма", "amount", 14),
    ("Телефон", "phone", 18),
]


def build_deliveries_export(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Доставки"

    for col_idx, (title, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=row["address"]).font = BODY_FONT
        ws.cell(row=row_idx, column=2, value=row["quantity"]).font = BODY_FONT
        amount_cell = ws.cell(row=row_idx, column=3, value=float(row["amount"]))
        amount_cell.font = BODY_FONT
        amount_cell.number_format = "#,##0.00"
        ws.cell(row=row_idx, column=4, value=row["phone"]).font = BODY_FONT

        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="right")

    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="Итого").font = Font(name="Arial", bold=True)
    total_qty = sum(r["quantity"] for r in rows)
    total_amount = sum(Decimal(str(r["amount"])) for r in rows)
    ws.cell(row=total_row, column=2, value=total_qty).font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=3, value=float(total_amount)).font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=3).number_format = "#,##0.00"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

